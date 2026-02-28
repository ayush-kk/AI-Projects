import logging
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import settings

logger = logging.getLogger(__name__)


def generate_excel(data: dict, user_id: int) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = data.get("sheet_name", "Sheet1")

    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell_font = Font(name="Calibri", size=11)
    cell_alignment = Alignment(vertical="center", wrap_text=True)

    title = data.get("title")
    start_row = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(data.get("headers", [])), 1))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Calibri", bold=True, size=16, color="4F46E5")
        title_cell.alignment = Alignment(horizontal="center")
        start_row = 3

    headers = data.get("headers", [])
    if headers:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        start_row += 1

    rows = data.get("rows", [])
    for row_idx, row in enumerate(rows):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=start_row + row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max(max_length + 4, 12), 50)

    output_dir = Path(settings.generated_dir) / str(user_id)
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = str(output_dir / filename)
    wb.save(filepath)
    return filepath
